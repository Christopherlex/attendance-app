import os
import io
from datetime import datetime, date, timezone, timedelta
from functools import wraps

import psycopg
from psycopg.rows import dict_row
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, jsonify, send_file, flash
)
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")

DATABASE_URL = os.environ.get("DATABASE_URL")

# WIB = UTC+7
WIB = timezone(timedelta(hours=7))

def now_wib():
    return datetime.now(WIB).replace(tzinfo=None)


# ─── DB helpers ───────────────────────────────────────────────────────────────

def get_db():
    conn = psycopg.connect(DATABASE_URL)
    return conn


def query(sql, params=(), fetchone=False, fetchall=False, commit=False):
    conn = get_db()
    try:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(sql, params)
            if commit:
                conn.commit()
                return cur.rowcount
            if fetchone:
                return cur.fetchone()
            if fetchall:
                return cur.fetchall()
    finally:
        conn.close()


# ─── Auth decorators ───────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            return redirect(url_for("worker_dashboard"))
        return f(*args, **kwargs)
    return decorated


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("admin_dashboard") if session["role"] == "admin" else url_for("worker_dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = query(
            "SELECT * FROM users WHERE username = %s AND password = %s",
            (username, password),
            fetchone=True
        )

        if user:
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["full_name"] = user["full_name"]
            session["role"] = user["role"]
            if user["role"] == "admin":
                return redirect(url_for("admin_dashboard"))
            else:
                return redirect(url_for("worker_dashboard"))
        else:
            error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── Worker ───────────────────────────────────────────────────────────────────

@app.route("/worker")
@login_required
def worker_dashboard():
    if session["role"] == "admin":
        return redirect(url_for("admin_dashboard"))

    open_session = query(
        "SELECT * FROM attendance WHERE user_id = %s AND checkout IS NULL ORDER BY checkin DESC LIMIT 1",
        (session["user_id"],),
        fetchone=True
    )
    return render_template("worker.html", open_session=open_session)


@app.route("/api/checkin", methods=["POST"])
@login_required
def checkin():
    if session["role"] == "admin":
        return jsonify({"error": "Admins cannot check in."}), 403

    data = request.get_json()
    location_name = (data.get("location_name") or "").strip()
    if not location_name:
        return jsonify({"error": "Location name is required."}), 400

    open_session = query(
        "SELECT id FROM attendance WHERE user_id = %s AND checkout IS NULL",
        (session["user_id"],),
        fetchone=True
    )
    if open_session:
        return jsonify({"error": "You already have an open check-in session."}), 400

    lat = data.get("lat")
    lng = data.get("lng")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr)
    ua = request.headers.get("User-Agent", "")
    checkin_time = now_wib()

    query(
        """INSERT INTO attendance
           (user_id, checkin, location_name, checkin_lat, checkin_lng, ip_address, user_agent)
           VALUES (%s, %s, %s, %s, %s, %s, %s)""",
        (session["user_id"], checkin_time, location_name, lat, lng, ip, ua),
        commit=True
    )
    return jsonify({"success": True, "message": f"Checked in at {location_name}."})


@app.route("/api/checkout", methods=["POST"])
@login_required
def checkout():
    if session["role"] == "admin":
        return jsonify({"error": "Admins cannot check out."}), 403

    data = request.get_json()
    open_session = query(
        "SELECT * FROM attendance WHERE user_id = %s AND checkout IS NULL ORDER BY checkin DESC LIMIT 1",
        (session["user_id"],),
        fetchone=True
    )
    if not open_session:
        return jsonify({"error": "No open check-in session found."}), 400

    lat = data.get("lat")
    lng = data.get("lng")
    checkout_time = now_wib()

    query(
        "UPDATE attendance SET checkout = %s, checkout_lat = %s, checkout_lng = %s WHERE id = %s",
        (checkout_time, lat, lng, open_session["id"]),
        commit=True
    )
    return jsonify({"success": True, "message": "Checked out successfully."})


@app.route("/api/status")
@login_required
def status():
    open_session = query(
        "SELECT * FROM attendance WHERE user_id = %s AND checkout IS NULL ORDER BY checkin DESC LIMIT 1",
        (session["user_id"],),
        fetchone=True
    )
    if open_session:
        return jsonify({
            "has_open": True,
            "location": open_session["location_name"],
            "checkin": open_session["checkin"].strftime("%Y-%m-%d %H:%M:%S") if open_session["checkin"] else None
        })
    return jsonify({"has_open": False})


# ─── Admin ────────────────────────────────────────────────────────────────────

@app.route("/admin")
@admin_required
def admin_dashboard():
    filter_month = request.args.get("month", "")   # format: "YYYY-MM"
    filter_worker = request.args.get("worker", "")

    all_workers = query(
        "SELECT full_name FROM users WHERE role = 'worker' ORDER BY full_name",
        fetchall=True
    )

    conditions = []
    params = []
    if filter_month:
        try:
            y, m = filter_month.split("-")
            conditions.append("EXTRACT(YEAR FROM a.checkin) = %s AND EXTRACT(MONTH FROM a.checkin) = %s")
            params.extend([int(y), int(m)])
        except ValueError:
            pass
    if filter_worker:
        conditions.append("u.full_name = %s")
        params.append(filter_worker)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    records = query(
        f"""SELECT a.*, u.full_name, u.username
            FROM attendance a
            JOIN users u ON u.id = a.user_id
            {where}
            ORDER BY a.checkin DESC""",
        params,
        fetchall=True
    ) or []

    return render_template("admin.html", records=records,
                           filter_month=filter_month, filter_worker=filter_worker,
                           all_workers=all_workers)


@app.route("/admin/export")
@admin_required
def export_excel():
    filter_month = request.args.get("month", "")
    filter_worker = request.args.get("worker", "")

    conditions = []
    params = []
    if filter_month:
        try:
            y, m = filter_month.split("-")
            conditions.append("EXTRACT(YEAR FROM a.checkin) = %s AND EXTRACT(MONTH FROM a.checkin) = %s")
            params.extend([int(y), int(m)])
        except ValueError:
            pass
    if filter_worker:
        conditions.append("u.full_name = %s")
        params.append(filter_worker)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    records = query(
        f"""SELECT a.*, u.full_name, u.username
            FROM attendance a
            JOIN users u ON u.id = a.user_id
            {where}
            ORDER BY a.checkin DESC""",
        params,
        fetchall=True
    ) or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["#", "Full Name", "Username", "Check-In (WIB)", "Check-Out (WIB)",
               "Duration", "Location", "Check-In GPS", "Check-Out GPS", "IP Address"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 18

    for i, r in enumerate(records, 1):
        checkin_str  = r["checkin"].strftime("%Y-%m-%d %H:%M:%S")  if r["checkin"]  else ""
        checkout_str = r["checkout"].strftime("%Y-%m-%d %H:%M:%S") if r["checkout"] else "—"

        if r["checkin"] and r["checkout"]:
            diff = r["checkout"] - r["checkin"]
            hours, rem = divmod(int(diff.total_seconds()), 3600)
            mins = rem // 60
            duration = f"{hours}h {mins}m"
        else:
            duration = "—"

        ci_gps = f"https://maps.google.com/?q={r['checkin_lat']},{r['checkin_lng']}"   if (r["checkin_lat"]  and r["checkin_lng"])  else ""
        co_gps = f"https://maps.google.com/?q={r['checkout_lat']},{r['checkout_lng']}" if (r["checkout_lat"] and r["checkout_lng"]) else ""

        row_data = [i, r["full_name"], r["username"], checkin_str, checkout_str,
                    duration, r["location_name"], ci_gps, co_gps, r["ip_address"]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = Alignment(horizontal="left")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4FA")

    label = filter_month or "all"
    if filter_worker:
        label += f"_{filter_worker.replace(' ', '_')}"
    fname = f"attendance_{label}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── DB init ──────────────────────────────────────────────────────────────────

@app.route("/setup")
def setup():
    token = request.args.get("token", "")
    if token != os.environ.get("SETUP_TOKEN", ""):
        return "Forbidden", 403

    conn = get_db()
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(80) UNIQUE NOT NULL,
                password TEXT NOT NULL,
                full_name TEXT NOT NULL,
                role VARCHAR(10) NOT NULL CHECK (role IN ('admin','worker'))
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id),
                checkin TIMESTAMP NOT NULL,
                checkout TIMESTAMP,
                location_name TEXT NOT NULL,
                checkin_lat FLOAT,
                checkin_lng FLOAT,
                checkout_lat FLOAT,
                checkout_lng FLOAT,
                ip_address TEXT,
                user_agent TEXT
            )
        """)
        conn.commit()
    conn.close()
    return "Tables created."


if __name__ == "__main__":
    app.run(debug=False)
